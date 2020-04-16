from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Color
from PIL import Image
from PIL import ImageGrab

xlsx=load_workbook('完成.xlsx')
worksheet = xlsx.active
pic = Image.open("1.png")
pic_width = pic.size[0]
pic_height = pic.size[1]
pix = pic.load()

for row in range(1, pic_height):
    for col in range(1, pic_width):
        cell = worksheet.cell(column=col+10, row=row+10)
        pixpoint = pix[col - 1, row - 1]
        pixcolor = "FF%02X%02X%02X" % (pixpoint[0], pixpoint[1], pixpoint[2])
        fill = PatternFill("solid", Color(rgb=pixcolor))
        cell.fill = fill
    worksheet.row_dimensions[row].height = 6
for col in range(1, pic_width):
    worksheet.column_dimensions[get_column_letter(col)].width = 1

img = ImageGrab.grab(bbox=(0, 0,1920,1080))
img.save("out.png")

xlsx.save('完成.xlsx')